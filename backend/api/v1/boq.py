from fastapi import APIRouter, HTTPException, status, UploadFile, File
from sqlalchemy import select, func
from pydantic import BaseModel
from typing import Optional, List
from uuid import UUID
import io

from core.deps import CurrentUser, DB
from models.core import BOQItem

router = APIRouter()


class BOQItemCreate(BaseModel):
    project_id: UUID
    parent_id: Optional[UUID] = None
    activity_id: Optional[UUID] = None
    item_number: Optional[str] = None
    description: str
    description_ar: Optional[str] = None
    unit: Optional[str] = None
    quantity: float = 0
    unit_rate: float = 0
    level: int = 1
    is_parent: bool = False
    sort_order: int = 0


class BOQItemUpdate(BaseModel):
    description: Optional[str] = None
    description_ar: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = None
    unit_rate: Optional[float] = None
    actual_quantity: Optional[float] = None
    activity_id: Optional[UUID] = None
    notes: Optional[str] = None


def _to_dict(b: BOQItem) -> dict:
    return {
        "id": str(b.id),
        "parent_id": str(b.parent_id) if b.parent_id else None,
        "activity_id": str(b.activity_id) if b.activity_id else None,
        "item_number": b.item_number,
        "description": b.description,
        "description_ar": b.description_ar,
        "unit": b.unit,
        "quantity": float(b.quantity or 0),
        "unit_rate": float(b.unit_rate or 0),
        "total_amount": float(b.quantity or 0) * float(b.unit_rate or 0),
        "actual_quantity": float(b.actual_quantity or 0),
        "level": b.level,
        "is_parent": b.is_parent,
        "sort_order": b.sort_order,
    }


@router.get("/")
async def list_boq(project_id: UUID, current_user: CurrentUser, db: DB):
    """Return full BOQ tree for a project, ordered for hierarchical display."""
    result = await db.execute(
        select(BOQItem)
        .where(BOQItem.project_id == project_id)
        .order_by(BOQItem.sort_order, BOQItem.item_number)
    )
    items = result.scalars().all()

    total_value = sum(float(i.quantity or 0) * float(i.unit_rate or 0) for i in items if not i.is_parent)

    return {
        "project_id": str(project_id),
        "total_value": total_value,
        "items_count": len(items),
        "items": [_to_dict(i) for i in items],
    }


@router.post("/", status_code=status.HTTP_201_CREATED)
async def create_boq_item(body: BOQItemCreate, current_user: CurrentUser, db: DB):
    item = BOQItem(**body.model_dump(exclude_none=True))
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return _to_dict(item)


@router.put("/{item_id}")
async def update_boq_item(item_id: UUID, body: BOQItemUpdate, current_user: CurrentUser, db: DB):
    result = await db.execute(select(BOQItem).where(BOQItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="BOQ item not found")

    for field, value in body.model_dump(exclude_none=True).items():
        setattr(item, field, value)
    await db.commit()
    await db.refresh(item)
    return _to_dict(item)


@router.delete("/{item_id}", status_code=204)
async def delete_boq_item(item_id: UUID, current_user: CurrentUser, db: DB):
    result = await db.execute(select(BOQItem).where(BOQItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="BOQ item not found")
    await db.delete(item)
    await db.commit()


@router.post("/import-excel")
async def import_boq_excel(
    project_id: UUID,
    file: UploadFile,
    current_user: CurrentUser,
    db: DB,
):
    """
    Import a BOQ from Excel. Expected columns (header row required):
    item_number | description | description_ar | unit | quantity | unit_rate
    """
    import openpyxl

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    required = ["description", "quantity", "unit_rate"]
    missing = [r for r in required if r not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}")

    col_idx = {h: i for i, h in enumerate(headers)}
    created = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            description = row[col_idx["description"]]
            if not description:
                continue
            quantity = float(row[col_idx["quantity"]] or 0)
            unit_rate = float(row[col_idx["unit_rate"]] or 0)

            item = BOQItem(
                project_id=project_id,
                item_number=str(row[col_idx.get("item_number", -1)]) if "item_number" in col_idx and row[col_idx["item_number"]] else None,
                description=str(description),
                description_ar=str(row[col_idx["description_ar"]]) if "description_ar" in col_idx and row[col_idx["description_ar"]] else None,
                unit=str(row[col_idx.get("unit", -1)]) if "unit" in col_idx and row[col_idx["unit"]] else None,
                quantity=quantity,
                unit_rate=unit_rate,
                sort_order=row_num,
            )
            db.add(item)
            created += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    await db.commit()
    return {"created": created, "errors": errors[:20], "total_errors": len(errors)}


@router.get("/export-excel")
async def export_boq_excel(project_id: UUID, current_user: CurrentUser, db: DB):
    """Export current BOQ to Excel — returns base64 file content."""
    import openpyxl
    import base64

    result = await db.execute(
        select(BOQItem).where(BOQItem.project_id == project_id).order_by(BOQItem.sort_order)
    )
    items = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.sheet_view.rightToLeft = True
    headers = ["م", "البند", "الوصف", "الوحدة", "الكمية", "سعر الوحدة", "الإجمالي"]
    ws.append(headers)

    for i, item in enumerate(items, 1):
        total = float(item.quantity or 0) * float(item.unit_rate or 0)
        ws.append([
            i, item.item_number or "", item.description_ar or item.description,
            item.unit or "", float(item.quantity or 0), float(item.unit_rate or 0), total,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    content = base64.b64encode(buf.getvalue()).decode()

    return {"filename": f"BOQ_{project_id}.xlsx", "content_base64": content}
